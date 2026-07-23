import os
import sys
from pathlib import Path

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# data_Step2.py 跨資料夾，所以需要額外動作來輔助 import
# 1. 取得目前檔案的 (Training.py) 所在目錄
current_dir = Path.cwd()

# 2. 取得上一層目錄 (專案的根目錄)
EfficientNet_B0 = os.path.dirname(current_dir)
src = os.path.dirname(EfficientNet_B0)
Project_Root = os.path.dirname(src)

# 3. 將根目錄加入系統搜尋路徑
sys.path.append(EfficientNet_B0)
sys.path.append(src)
sys.path.append(Project_Root)

# 4. 開始 import

import torch
torch.autograd.set_detect_anomaly(True)
import torch.nn as nn
import torch.optim as optim
import matplotlib.pyplot as plt
import matplotlib as mpl
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from torch.utils.data import DataLoader
from tqdm import tqdm
from src.data_Step2_5_Test import get_test_dataset
from src.data_Step2_5_Train import get_train_dataset
from src.EfficientNet_B0.EfficientNet_B0_GeM import build_model

# ---------------------------------
# 1. 設定參數與裝置
# ---------------------------------
BATCH_SIZE = 32 # 根據顯卡記憶體調整 (16 或 32)
Learning_Rate = 1e-4 # Adam 的標準學習率
Num_Epoch = 200
IMG_WIDTH = 512
IMG_HEIGHT = 128
DEVICE = torch.device('cuda')
TRAIN_DIR = Project_Root + '/Datasets/Dataset_Step1'
TEST_ROOT = Project_Root + '/Datasets/Dataset_Step2'
FIG_DIR = Project_Root + '/Figures/'
XLSX_DIR = Project_Root + '/Figures/'
MODEL_PATH = 'EfficientNet_B0_GeM_model.pth'
DA_ACCURACY = {}

TEST_DATALOADER_DICT = {}
TRAIN_DATALOADER = None

def define_data_loaders():
    global TRAIN_DATALOADER, TEST_DATALOADER_DICT, TRAIN_DIR, IMG_WIDTH, IMG_HEIGHT, DEVICE, BATCH_SIZE
    train_dataset = get_train_dataset(TRAIN_DIR, IMG_WIDTH, IMG_HEIGHT, is_train=True)
    TRAIN_DATALOADER = DataLoader(train_dataset
               , batch_size=BATCH_SIZE
               , shuffle=True)

    da_conditions = {'Brightness': list(range(0, 210, 10))}
    for category, values in da_conditions.items():
        TEST_DATALOADER_DICT[category] = {}
        DA_ACCURACY[category] = {}
        for val in values:
            DA_ACCURACY[category][val] = list()
            test_dir = TEST_ROOT + f'/{category}/{val}%'
            test_dataset = get_test_dataset(test_dir)

            TEST_DATALOADER_DICT[category][val] = DataLoader(test_dataset
                                     , batch_size=BATCH_SIZE
                                     , shuffle=False)


def model_test(model):
    global TEST_DATALOADER_DICT, DA_ACCURACY
    model.eval()
    for category, val_dict in TEST_DATALOADER_DICT.items():
        for val, test_loader in val_dict.items():
            correct = 0
            total = 0
            with torch.no_grad():
                for images, labels in tqdm(test_loader, desc="Testing", unit='batch'):
                    images, labels = images.to(DEVICE), labels.to(DEVICE)

                    outputs = model(images)
                    _, predicted = torch.max(outputs, 1)

                    total += labels.size(0)
                    correct += (predicted == labels).sum().item()
            DA_ACCURACY['Brightness'][val].append(100 * correct / total)

def model_training_and_test ():
    global TRAIN_DATALOADER, DEVICE, MODEL_PATH, Num_Epoch
    # ---------------------------------------------------
    # 3. 初始化模型
    # ---------------------------------------------------
    model = build_model(num_classes=len(TRAIN_DATALOADER.dataset.classes)).to(DEVICE)
    loss_func = nn.CrossEntropyLoss().to(DEVICE)
    optimizer = optim.Adam(model.parameters(), lr=Learning_Rate)
    scheduler = torch.optim.lr_scheduler.ReduceLROnPlateau(
        optimizer,mode='min', factor=0.5,patience = 4,min_lr = 1e-6
    )

    # ---------------------------------------------------
    # 4. 開始訓練
    # ---------------------------------------------------
    best_loss = float('inf')

    if os.path.exists(MODEL_PATH):
        checkpoint = torch.load(MODEL_PATH)
        model.load_state_dict(checkpoint['model_state_dict'])
        best_loss = checkpoint['best_loss']


    for epoch in range(Num_Epoch):
        print(f"--> 開始訓練...{epoch+1}/{Num_Epoch}")
        model.train()
        running_loss = 0.0

        with tqdm(TRAIN_DATALOADER, ncols=100) as loop:
            for images, labels in loop:
                images, labels = images.to(DEVICE), labels.to(DEVICE)

                optimizer.zero_grad()
                outputs = model(images)
                loss = loss_func(outputs, labels)
                loss.backward()
                optimizer.step()

                running_loss += loss.item()
                loop.set_postfix(loss=loss.item())
                current_lr = optimizer.param_groups[0]['lr']

        avg_loss = running_loss / len(TRAIN_DATALOADER)
        scheduler.step(avg_loss)

        print(f"Loss: {avg_loss:.4f} | LR: {current_lr:.8f}")


        if avg_loss < best_loss:
            best_loss = avg_loss

            checkpoint = {
                'model_state_dict': model.state_dict(),
                'best_loss': best_loss,
            }

            torch.save(checkpoint, MODEL_PATH)
        model_test(model)

def generate_meeting_figure():
    global DA_ACCURACY,FIG_DIR
    epochs = np.arange(1, Num_Epoch+1)
    # 全局設置
    plt.rcParams.update({
        'font.family': 'sans-serif',  # 使用非襯線字體，閱讀性較佳
        'font.serif': ['Arial', 'Helvetica', 'DejaVu Sans'],
        'font.size': 18,                 # 基準字體大小
        'axes.labelsize': 20,            # 軸標籤字體加大
        'axes.titlesize': 24,            # 標題字體最大
        'xtick.labelsize': 16,           # X軸刻度字體放大
        'ytick.labelsize': 16,           # Y軸刻度字體放大
        'legend.fontsize': 16,           # 圖例字體放大
        'legend.frameon': True,          # 簡報中圖例建議加上外框，避免與粗線條數據混淆
        'lines.linewidth': 3.0,          # 線條寬度加粗
        'axes.linewidth': 1.5,           # 座標軸線框加粗
        'grid.linestyle': '--',          # 網格線樣式
        'grid.alpha': 0.3,               # 網格線透明度 (調淡避免干擾數據)
        'figure.dpi': 300,               # 論文要求的高解析度基準
        'axes.spines.top': False,        # 隱藏右方邊框
        'axes.spines.right': False       # 隱藏右方邊框
    })
    # 設定連續色階
    # 顏色映射基準：最小值 0，最大值 200
    cmap = plt.get_cmap('viridis')
    norm = mpl.colors.Normalize(vmin=0, vmax=200)
    fig, ax = plt.subplots(figsize=(10, 5.625))
    for category, values in DA_ACCURACY.items():
        for val,accuracy in values.items():
            line_color = cmap(norm(val))
            ax.plot(epochs, accuracy, color=line_color, linestyle='-')
    ax.set_xlabel("Epochs")
    ax.set_ylabel("Accuracy (%)")
    ax.set_title(f'Accuracy Validation')
    ax.set_ylim(0, 100)

    # ---- 建立 Colorbar ---- #

    # 建立一個 ScalarMappable 物件供 Colorbar 使用
    sm = plt.cm.ScalarMappable(cmap=cmap, norm=norm)
    sm.set_array([])

    # 在右側加上 Colorbar，pad 參數控制與主圖的距離
    cbar = fig.colorbar(sm, ax=ax, pad=0.03)

    # 設定 Colorbar 的標籤與字體大小
    cbar.set_label('DA Intensity (%)')

    # 設定只顯示 0%, 50%, 100%, 150%, 200% 的刻度
    cbar.set_ticks(np.arange(0, 210, 50))
    cbar.set_ticklabels(np.arange(0, 210, 50))


    # ===========================
    # 6. 最後調整與存檔
    # ===========================

    # 儲存圖片
    # dpi=300 是印刷品質的標準
    # bbox_inches='tight' 確保儲存時去除多餘白邊
    plt.tight_layout()
    fig_name = f"EfficientNet_GeM_Accuracy.png"
    save_path = Path(FIG_DIR) / "EfficientNet_Advanced" / f"{fig_name}"
    save_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(save_path, bbox_inches='tight')
    plt.close()

def generate_paper_figure():
    global DA_ACCURACY,FIG_DIR
    epochs = np.arange(1, Num_Epoch+1)
    # 全局設置
    plt.rcParams.update({
        'font.family': 'serif',  # 使用襯線字體
        'font.serif': ['Times New Roman'],
        'font.size': 10,                 # 基準字體大小 (對齊論文正文 10pt)
        'axes.labelsize': 10,            # 軸標籤字體大小 (不應大於正文)
        'axes.titlesize': 10,            # 標題字體大小
        'xtick.labelsize': 8,            # X軸刻度字體大小 (略小於正文)
        'ytick.labelsize': 8,            # Y軸刻度字體大小 (略小於正文)
        'legend.fontsize': 8,            # 圖例字體大小
        'legend.frameon': False,         # 論文圖例通常不加上外框，保持乾淨
        'lines.linewidth': 1.2,          # 線條寬度 (縮小為 1.2，避免在 3.5 吋畫布上太粗)
        'axes.linewidth': 0.8,           # 座標軸線框寬度
        'grid.linestyle': '--',          # 網格線樣式
        'grid.alpha': 0.5,               # 網格線透明度 (調淡避免干擾數據)
        'figure.dpi': 300,               # 論文要求的高解析度基準
        'axes.spines.top': False,        # 隱藏右方邊框
        'axes.spines.right': False       # 隱藏右方邊框
    })
    color = '#d62728' # 紅
    for category, values in DA_ACCURACY.items():
        for val,accuracy in values.items():
            fig, ax = plt.subplots(figsize=(3.5, 2.5))
            ax.plot(epochs, accuracy, label="Model", color=color, linestyle='-')
            ax.set_xlabel("Epochs")
            ax.set_ylabel("Accuracy (%)")
            ax.set_title(f'Accuracy Validation: {category} {val}%')
            ax.legend(loc='lower right')
            ax.set_ylim(0, 100)

            # ===========================
            # 6. 最後調整與存檔
            # ===========================

            # 儲存圖片
            # dpi=300 是印刷品質的標準
            # bbox_inches='tight' 確保儲存時去除多餘白邊
            plt.tight_layout()

            fig_name = f"{category}"
            if category == 'Horizontal_Roll':
                fig_name += f"_{val}°"
            elif category != 'Origin':
                fig_name += f"_{val}%"
            fig_name += '_Accuracy.png'
            save_path = Path(FIG_DIR) / "EfficientNet_Advanced"/ "GeM" / f"{category}" / f"{fig_name}"
            save_path.parent.mkdir(parents=True, exist_ok=True)
            fig.savefig(save_path, bbox_inches='tight')
            plt.close()

def gernerate_xlsx():
    global DA_ACCURACY
    # ==========================================
    # 階段一：準備資料與 Pandas 處理
    # ==========================================

    formatted_data = {}
    for category, values in DA_ACCURACY.items():
        for val,accuracy in values.items():
            formatted_data[(category, val)] = accuracy

    df = pd.DataFrame(formatted_data)
    df.index = range(1, len(df) + 1)
    df.index.name = 'Epoch'
    df.columns.names = ['DA Topic', 'Intensity']
    file_path = Path(XLSX_DIR) / "EfficientNet_Advanced"/ "GeM" / "DA_Accuracy_Final.xlsx"
    df.to_excel(file_path, engine='openpyxl')

    # ==========================================
    # 階段二：設定 Openpyxl 的論文級樣式
    # ==========================================
    # 論文標準字體設定
    font_header = Font(name="Times New Roman", size=11, bold=True, color="FFFFFF")
    font_index = Font(name="Times New Roman", size=11, bold=True)
    font_regular = Font(name="Times New Roman", size=11)

    # 顏色設定：學術藍底、斑馬紋
    fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F6F9", end_color="F2F6F9", fill_type="solid")

    # 對齊與邊框
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    thin_side = Side(border_style='thin', color='D3D3D3')
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # ==========================================
    # 階段三：套用樣式並儲存
    # ==========================================
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    ws.title = "Training Logs"

    # 取得目前資料的最大行與列
    max_row = ws.max_row
    max_col = ws.max_column

    # 有訪每一個儲存格套用樣式
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border_all

            # Pandas 產生的雙層表頭會佔用前 3 列 (Row 1~3)
            if r <= 3:
                # 處理表頭區塊
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
            else:
                # 處理數據區塊 (Row 4 到 Row 203)
                cell.font = font_regular

                # 第一行 (Epoch 數字)
                if c == 1:
                    cell.font = font_index
                    cell.alignment = align_center
                # 其他行 (Accuracy 數據)
                else:
                    cell.alignment = align_right
                    cell.number_format = "0.00"  # 準確率保留兩位小數

                # 偶數列套用斑馬紋底色以利閱讀
                if r % 2 == 0:
                    cell.fill = fill_zebra

    # 凍結窗格：向下捲動時鎖定表頭 (Row 1~3)，向右捲動時鎖定 Epoch (Col A)
    ws.freeze_panes = "B4"

    # 調整 Epoch 欄位的寬度
    ws.column_dimensions['A'].width = 10
    # 動態調整其他數據欄位的寬度
    for col_idx in range(2, max_col + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 14

    wb.save(file_path)
def main():
    define_data_loaders()
    model_training_and_test()
    generate_paper_figure()
    generate_meeting_figure()
    gernerate_xlsx()
if __name__ == '__main__':
    main()