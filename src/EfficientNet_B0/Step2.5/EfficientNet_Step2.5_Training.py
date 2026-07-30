import os
import sys
import torch
import torch.nn as nn
import torch.optim as optim
import numpy as np
from torch.utils.data import DataLoader
import matplotlib.pyplot as plt
from tqdm import tqdm
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 允許 KMP 重複初始化（防止部分環境下 OpenMP 報錯）
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# ----------------- 精準路徑設定（完全對齊專案目錄結構） -----------------
current_dir = os.path.dirname(os.path.abspath(__file__))      # 當前資料夾 (例如 Step2.5)
EfficientNet_B0_dir = os.path.dirname(current_dir)             # 上一層模型資料夾
src = os.path.dirname(EfficientNet_B0_dir)                     # 專案的 src 原始碼目錄
Project_Root = os.path.dirname(src)                           # 專案根目錄

# 將環境變數路徑依序加入，防範模組匯入找不到路徑
sys.path.append(current_dir)
sys.path.append(EfficientNet_B0_dir)
sys.path.append(src)
sys.path.append(Project_Root)

# 匯入資料集讀取函式
from src.data_Step2_5_Train import get_train_dataset
from src.data_Step2_5_Test import get_test_dataset
from src.EfficientNet_B0.EfficientNet_B0_modified_version import build_model


Learning_Rate = 1e-4
IMG_WIDTH = 512
IMG_HEIGHT = 128
BATCH_SIZE = 32
epochs = 200



# 權重與圖表儲存路徑變更 (對應 EfficientNet-B0 專屬資料夾)
MODEL_PATH = 'Step2.5_efficientnet_B0_model.pth'
FIG_DIR = Path(f"{Project_Root}/Figures")
EXCEL_DIR = Path(f"{Project_Root}/Figures")

GLOBAL_VAL_LOADERS = {}


def get_val_dataloader(da_type, da_value):
    # 完全對齊 MobileNet 的驗證集路徑分支與 Dataset 讀取邏輯
    if da_type == 'Origin':
        img_path = os.path.join(Project_Root, "Datasets/Dataset_Step1")
        dataset = get_train_dataset(root_dir=img_path, width=IMG_WIDTH, height=IMG_HEIGHT, is_train=False)
    elif da_type == 'Horizontal_Roll':
        img_path = os.path.join(Project_Root, "Datasets/Dataset_Step2", da_type, f"{da_value}°")
        dataset = get_test_dataset(root_dir=img_path)
    else:
        img_path = os.path.join(Project_Root, "Datasets/Dataset_Step2", da_type, f"{da_value}%")
        dataset = get_test_dataset(root_dir=img_path)

    val_loader = DataLoader(
        dataset=dataset,
        batch_size=BATCH_SIZE,
        shuffle=False,
    )
    return val_loader


def setup_val_dataloaders(da_conditions):
    global GLOBAL_VAL_LOADERS
    for da_type, val_list in da_conditions.items():
        GLOBAL_VAL_LOADERS[da_type] = {}
        print(f"Building DataLoaders for {da_type}...")
        for da_val in val_list:
            GLOBAL_VAL_LOADERS[da_type][da_val] = get_val_dataloader(da_type, da_val)


def run_inference(model, device, dataloader, da_type, da_val):
    model.eval()
    correct = 0
    total = 0

    with torch.no_grad():
        for img, id_label in tqdm(dataloader, desc=f"Val: {da_type} {da_val}", leave=False, ncols=100):
            img, id_label = img.to(device), id_label.to(device)
            outputs = model(img)
            _, predicted = torch.max(outputs.data, 1)
            total += id_label.size(0)
            correct += (predicted == id_label).sum().item()

    if total == 0:
        return 0.0
    return (correct / total) * 100.0


def paper_plot_and_save_curves(da_history, current_epoch):
    plt.rcParams.update({
        'font.family': 'serif',
        'font.serif': ['Times New Roman'],
        'font.size': 10,
        'axes.labelsize': 10,
        'axes.titlesize': 10,
        'xtick.labelsize': 8,
        'ytick.labelsize': 8,
        'legend.fontsize': 8,
        'legend.frameon': False,
        'lines.linewidth': 1.2,
        'axes.linewidth': 0.8,
        'grid.linestyle': '--',
        'grid.alpha': 0.5,
        'figure.dpi': 300,
        'axes.spines.top': False,
        'axes.spines.right': False
    })

    for da_type, values_dict in da_history.items():

        for val, acc_list in values_dict.items():
            if da_type == 'Horizontal_Roll':
                unit = '°'
            elif da_type == 'Origin':
                unit = ''
            else:
                unit = '%'

            fig, ax = plt.subplots(figsize=(3.5, 2.5))
            ax.plot(range(1, len(acc_list) + 1), acc_list, label=f'Model', color='#d62728')
            ax.set_xlabel('Epoch')
            ax.set_ylabel('Accuracy (%)')
            ax.set_title(f'Accuracy Validation: {da_type} {val}{unit}')
            ax.set_ylim(0, 100)

            if len(acc_list) > 0:
                ax.legend(loc='lower right')

            filename = f"{da_type}_{val}{unit}.png"
            sub_dir = FIG_DIR / "Step2.5" / "EfficientNet-B0" / da_type / filename
            sub_dir.parent.mkdir(parents=True, exist_ok=True)
            fig.savefig(sub_dir, bbox_inches='tight')
            if da_type == "Brightness":
                sub_dir = FIG_DIR / "EfficientNet_Advanced" / "Baseline" / da_type / filename
                sub_dir.parent.mkdir(parents=True, exist_ok=True)
                fig.savefig(sub_dir, bbox_inches='tight')
            plt.close(fig)

def meeting_plot_and_save_curves(da_history, current_epoch):
    plt.rcParams.update({
        'font.family': 'sans-serif',
        'font.serif': ['Arial', 'Helvetica', 'DejaVu Sans'],
        'font.size': 18,
        'axes.labelsize': 20,
        'axes.titlesize': 24,
        'xtick.labelsize': 16,
        'ytick.labelsize': 16,
        'legend.fontsize': 16,
        'legend.frameon': True,
        'lines.linewidth': 3.0,
        'axes.linewidth': 1.5,
        'grid.linestyle': '--',
        'grid.alpha': 0.3,
        'figure.dpi': 300,
        'axes.spines.top': False,
        'axes.spines.right': False
    })

    for da_type, values_dict in da_history.items():
        if da_type == 'Brightness':
            fig, ax = plt.subplots(figsize=(10, 5.625))
            cmap = plt.cm.viridis
            norm = plt.Normalize(vmin=0, vmax=200)
            for val, acc_list in values_dict.items():
                if len(acc_list) > 0:
                    line_color = cmap(norm(val))
                    ax.plot(range(1, len(acc_list) + 1), acc_list, color=line_color, linestyle='-')
            ax.set_xlabel('Epoch')
            ax.set_ylabel('Accuracy (%)')
            ax.set_title(f'Accuracy Validation')
            ax.set_ylim(0, 100)

            sm = plt.cm.ScalarMappable(cmap=cmap, norm=norm)
            sm.set_array([])
            cbar = fig.colorbar(sm, ax=ax, pad=0.03)
            cbar.set_label('DA Intensity (%)')
            cbar.set_ticks(np.arange(0, 210, 50))
            cbar.set_ticklabels(np.arange(0, 210, 50))

            filename = f"EfficientNet_B0_Baseline_Accuracy.png"
            sub_dir = FIG_DIR / "EfficientNet_Advanced" / "Baseline" / filename
            sub_dir.parent.mkdir(parents=True, exist_ok=True)
            fig.savefig(sub_dir, bbox_inches='tight')
            plt.close(fig)

def save_to_paper_excel(da_history, filename='DA_Accuracy_History.xlsx'):
    filepath = Path(EXCEL_DIR)
    filepath = filepath / "Step2.5" / "EfficientNet-B0" / filename
    filepath.parent.mkdir(parents=True, exist_ok=True)
    formatted_data = {}
    for da_type, values in da_history.items():
        for val, acc_list in values.items():
            formatted_data[(da_type, val)] = acc_list

    df = pd.DataFrame(formatted_data)
    df.index = range(1, len(df) + 1)
    df.index.name = 'Epoch'
    df.columns.names = ['DA Topic', 'Intensity']
    df.to_excel(filepath, engine='openpyxl')

    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        ws.title = "Training Logs"

        font_header = Font(name="Times New Roman", size=11, bold=True, color="FFFFFF")
        font_index = Font(name="Times New Roman", size=11, bold=True)
        font_regular = Font(name="Times New Roman", size=11)

        fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        fill_zebra = PatternFill(start_color="F2F6F9", end_color="F2F6F9", fill_type="solid")

        align_center = Alignment(horizontal='center', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')
        thin_side = Side(border_style='thin', color='D3D3D3')
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        max_row = ws.max_row
        max_col = ws.max_column

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border_all

                if r <= 3:
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = align_center
                else:
                    cell.font = font_regular
                    if c == 1:
                        cell.font = font_index
                        cell.alignment = align_center
                    else:
                        cell.alignment = align_right
                        cell.number_format = "0.00"

                    if r % 2 == 0:
                        cell.fill = fill_zebra

        ws.freeze_panes = "B4"
        ws.column_dimensions['A'].width = 10
        for col_idx in range(2, max_col + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 14

        wb.save(filepath)
    except Exception:
        pass

def save_to_meeting_excel(da_history, filename='DA_Accuracy_History.xlsx'):
    filepath = Path(EXCEL_DIR)
    filepath = filepath / "EfficientNet_Advanced" / "Baseline" / filename
    filepath.parent.mkdir(parents=True, exist_ok=True)
    formatted_data = {}
    for da_type, values in da_history.items():
        for val, acc_list in values.items():
            if da_type == 'Brightness':
                formatted_data[(da_type, val)] = acc_list

    df = pd.DataFrame(formatted_data)
    df.index = range(1, len(df) + 1)
    df.index.name = 'Epoch'
    df.columns.names = ['DA Topic', 'Intensity']
    df.to_excel(filepath, engine='openpyxl')

    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        ws.title = "Training Logs"

        font_header = Font(name="Times New Roman", size=11, bold=True, color="FFFFFF")
        font_index = Font(name="Times New Roman", size=11, bold=True)
        font_regular = Font(name="Times New Roman", size=11)

        fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        fill_zebra = PatternFill(start_color="F2F6F9", end_color="F2F6F9", fill_type="solid")

        align_center = Alignment(horizontal='center', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')
        thin_side = Side(border_style='thin', color='D3D3D3')
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        max_row = ws.max_row
        max_col = ws.max_column

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border_all

                if r <= 3:
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = align_center
                else:
                    cell.font = font_regular
                    if c == 1:
                        cell.font = font_index
                        cell.alignment = align_center
                    else:
                        cell.alignment = align_right
                        cell.number_format = "0.00"

                    if r % 2 == 0:
                        cell.fill = fill_zebra

        ws.freeze_panes = "B4"
        ws.column_dimensions['A'].width = 10
        for col_idx in range(2, max_col + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 14

        wb.save(filepath)
    except Exception:
        pass


def Efficientnet_training():
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    img_path = os.path.join(Project_Root, "Datasets/Dataset_Step1")
    dataset = get_train_dataset(root_dir=img_path, width=IMG_WIDTH, height=IMG_HEIGHT, is_train=True)

    trainloader = DataLoader(
        dataset=dataset,
        batch_size=BATCH_SIZE,
        shuffle=True,
    )

    # 建立改版後的 EfficientNet_B0 模型
    model = build_model(num_classes=1000)
    model = model.to(device)
    criterion = nn.CrossEntropyLoss()
    optimizer = optim.Adam(model.parameters(), lr=Learning_Rate)
    scheduler = torch.optim.lr_scheduler.ReduceLROnPlateau(
        optimizer, mode='min', factor=0.5, patience=10, min_lr=1e-6
    )

    # 定義 90 種數據增強驗證條件
    da_conditions = {
        'Origin': ['Baseline'],
        'Brightness': list(range(0, 210, 10)),
        'ColorTemperature': list(range(0, 210, 10)),
        'Grid_Mask': list(range(0, 110, 10)),
        'Horizontal_Roll': list(range(0, 370, 10))
    }

    da_history = {da: {val: [] for val in vals} for da, vals in da_conditions.items()}

    print("Pre-building all Validation DataLoaders (Level 2)...")
    setup_val_dataloaders(da_conditions)

    start_epoch = 0
    best_loss = float('inf')
    epoch_losses = []

    # 權重續訓機制與歷史數據恢復
    if os.path.exists(MODEL_PATH):
        try:
            print(f"Loading weights from {MODEL_PATH}")
            checkpoint = torch.load(MODEL_PATH, map_location=device)
            model.load_state_dict(checkpoint['model_state_dict'])

            # 嘗試載入優化器與排程器
            if 'optimizer_state_dict' in checkpoint:
                optimizer.load_state_dict(checkpoint['optimizer_state_dict'])
            if 'scheduler_state_dict' in checkpoint:
                scheduler.load_state_dict(checkpoint['scheduler_state_dict'])

            best_loss = checkpoint['best_loss']
            start_epoch = checkpoint['epoch']
            print("Weight loaded successfully.")

            excel_path = Path(EXCEL_DIR) / "Step2.5" / "EfficientNet-B0" / 'DA_Accuracy_History.xlsx'
            if os.path.exists(excel_path):
                try:
                    print("Loading historical data from Excel for continuous plotting...")
                    df_history = pd.read_excel(excel_path, header=[0, 1], index_col=0)
                    for da_type in da_conditions.keys():
                        for val in da_conditions[da_type]:
                            col_tuple_int = (da_type, val)
                            col_tuple_str = (da_type, str(val))

                            if col_tuple_int in df_history.columns:
                                da_history[da_type][val] = df_history[col_tuple_int].dropna().tolist()[:start_epoch]
                            elif col_tuple_str in df_history.columns:
                                da_history[da_type][val] = df_history[col_tuple_str].dropna().tolist()[:start_epoch]

                    baseline_len = len(da_history['Origin']['Baseline'])
                    print(f"Plot history recovered successfully! Current recorded epochs: {baseline_len}")
                except Exception as e:
                    print(f"Failed to recover plot history: {e}")
            # ==================================

        except Exception as e:
            print(f"Loading failed: {e}, training from scratch.")
    else:
        print("No existing weights found. Training from scratch.")

    print("Start Training...")

    for epoch in range(start_epoch, epochs):
        model.train()
        running_loss = 0.0

        with tqdm(trainloader, desc=f"Epoch {epoch + 1}/{epochs}", ncols=100, leave=True) as loop:
            for img, id_label in loop:
                img, id_label = img.to(device), id_label.to(device)

                optimizer.zero_grad()
                outputs = model(img)
                loss = criterion(outputs, id_label)
                loss.backward()

                optimizer.step()

                running_loss += loss.item()
                current_lr = optimizer.param_groups[0]['lr']
                loop.set_postfix(loss=f"{loss.item():.4f}")

        avg_loss = running_loss / len(trainloader)
        epoch_losses.append(avg_loss)
        scheduler.step(avg_loss)

        print(f"Epoch {epoch + 1} | Loss: {avg_loss:.4f} | LR: {current_lr:.8f}")

        # 執行 90 種條件的推理與精確度記錄
        print(f"Running Inference for 90 DA conditions...")
        for da_type, val_list in da_conditions.items():
            for da_val in val_list:
                val_loader = GLOBAL_VAL_LOADERS[da_type][da_val]
                accuracy = run_inference(model, device, val_loader, da_type, da_val)
                da_history[da_type][da_val].append(accuracy)

        # 繪圖並導出至 Excel 報表
        paper_plot_and_save_curves(da_history, epoch)
        meeting_plot_and_save_curves(da_history, epoch)

        save_to_paper_excel(da_history)
        save_to_meeting_excel(da_history)

        # 儲存最優模型
        if avg_loss < best_loss:
            best_loss = avg_loss
            checkpoint = {
                'model_state_dict': model.state_dict(),
                'optimizer_state_dict': optimizer.state_dict(),  # 新增：儲存優化器動量
                'scheduler_state_dict': scheduler.state_dict(),  # 新增：儲存排程器紀錄
                'best_loss': best_loss,
                'epoch': epoch + 1
            }
            torch.save(checkpoint, MODEL_PATH)
            print(f"New Best Model Saved to {MODEL_PATH}")

    return epoch_losses


def main():
    Efficientnet_training()

if __name__ == '__main__':
    main()