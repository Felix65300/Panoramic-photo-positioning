import os
import sys
import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import DataLoader
import matplotlib.pyplot as plt
from tqdm import tqdm
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

current_dir = os.path.dirname(os.path.abspath(__file__))
Mobilenet_V3_small = os.path.dirname(current_dir)
src = os.path.dirname(Mobilenet_V3_small)
Project_Root = os.path.dirname(src)

sys.path.append(Mobilenet_V3_small)
sys.path.append(src)
sys.path.append(Project_Root)

from src.data_Step2_5_Train import get_train_dataset
from src.Mobilenet_V3_small.Mobilenet_V3_small_modified_version import build_model

Learning_Rate = 1e-4
IMG_WIDTH = 512
IMG_HEIGHT = 128
BATCH_SIZE = 32
epochs = 200

MODEL_PATH = 'Step2.5_mobilenet_model.pth'
FIG_DIR = os.path.join(Project_Root, 'Figures', 'Step2.5', 'MobileNet-V3-Small')

GLOBAL_VAL_LOADERS = {}


def get_val_dataloader(da_type, da_value):
    if da_type == 'Origin':
        img_path = os.path.join(Project_Root, "Datasets/Dataset_Step1")
    elif da_type == 'Horizontal_Roll':
        img_path = os.path.join(Project_Root, "Datasets/Dataset_Step2", da_type, f"{da_value}°")
    else:
        img_path = os.path.join(Project_Root, "Datasets/Dataset_Step2", da_type, f"{da_value}%")

    dataset = get_train_dataset(root_dir=img_path, width=IMG_WIDTH, height=IMG_HEIGHT, is_train=False)

    val_loader = DataLoader(
        dataset=dataset,
        batch_size=BATCH_SIZE,
        shuffle=False,
    )
    return val_loader


def setup_val_dataloaders(da_conditions):
    global GLOBAL_VAL_LOADERS
    for da_type, val_list in da_conditions.items():
        GLOBAL_VAL_LOADERS[da_type] = {}
        print(f"Building DataLoaders for {da_type}...")
        for da_val in val_list:
            GLOBAL_VAL_LOADERS[da_type][da_val] = get_val_dataloader(da_type, da_val)


def run_inference(model, device, dataloader, da_type, da_val):
    model.eval()
    correct = 0
    total = 0

    with torch.no_grad():
        for img, id_label in tqdm(dataloader, desc=f"Val: {da_type} {da_val}", leave=False, ncols=100):
            img, id_label = img.to(device), id_label.to(device)
            outputs = model(img)
            _, predicted = torch.max(outputs.data, 1)
            total += id_label.size(0)
            correct += (predicted == id_label).sum().item()

    if total == 0:
        return 0.0
    return (correct / total) * 100.0


def plot_and_save_curves(da_history, current_epoch):
    plt.rcParams.update({
        'font.family': 'serif',
        'font.serif': ['Times New Roman'],
        'font.size': 10,
        'axes.labelsize': 10,
        'axes.titlesize': 10,
        'xtick.labelsize': 8,
        'ytick.labelsize': 8,
        'legend.fontsize': 8,
        'legend.frameon': False,
        'lines.linewidth': 1.2,
        'axes.linewidth': 0.8,
        'grid.linestyle': '--',
        'grid.alpha': 0.5,
        'figure.dpi': 300
    })

    for da_type, values_dict in da_history.items():
        sub_dir = os.path.join(FIG_DIR, da_type)
        os.makedirs(sub_dir, exist_ok=True)

        for val, acc_list in values_dict.items():
            if da_type == 'Horizontal_Roll':
                unit = '°'
            elif da_type == 'Origin':
                unit = ''
            else:
                unit = '%'

            fig, ax = plt.subplots(figsize=(3.5, 2.5))
            ax.plot(range(1, current_epoch + 2), acc_list, label=f'{da_type} {val}{unit}', color='#d62728')
            ax.set_xlabel('Epoch')
            ax.set_ylabel('Accuracy (%)')
            ax.set_title(f'Accuracy Validation: {da_type} {val}{unit}')
            ax.set_xlim(1, epochs)
            ax.set_ylim(0, 105)
            ax.grid(True)

            if len(acc_list) > 0:
                ax.legend()

            filename = f"{da_type}_{val}{unit}.png"
            fig.savefig(os.path.join(sub_dir, filename), bbox_inches='tight')
            plt.close(fig)


def save_to_excel(da_history, filename='DA_Accuracy_History.xlsx'):
    formatted_data = {}
    for da_type, values in da_history.items():
        for val, acc_list in values.items():
            formatted_data[(da_type, val)] = acc_list

    df = pd.DataFrame(formatted_data)
    df.index = range(1, len(df) + 1)
    df.index.name = 'Epoch'
    df.columns.names = ['DA Topic', 'Intensity']
    df.to_excel(filename, engine='openpyxl')

    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        ws.title = "Training Logs"

        font_header = Font(name="Times New Roman", size=11, bold=True, color="FFFFFF")
        font_index = Font(name="Times New Roman", size=11, bold=True)
        font_regular = Font(name="Times New Roman", size=11)

        fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        fill_zebra = PatternFill(start_color="F2F6F9", end_color="F2F6F9", fill_type="solid")

        align_center = Alignment(horizontal='center', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')
        thin_side = Side(border_style='thin', color='D3D3D3')
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        max_row = ws.max_row
        max_col = ws.max_column

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border_all

                if r <= 3:
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = align_center
                else:
                    cell.font = font_regular
                    if c == 1:
                        cell.font = font_index
                        cell.alignment = align_center
                    else:
                        cell.alignment = align_right
                        cell.number_format = "0.00"

                    if r % 2 == 0:
                        cell.fill = fill_zebra

        ws.freeze_panes = "B4"
        ws.column_dimensions['A'].width = 10
        for col_idx in range(2, max_col + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 14

        wb.save(filename)
    except Exception:
        pass


def Mobilenet_training():
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    img_path = os.path.join(Project_Root, "Datasets/Dataset_Step1")

    dataset = get_train_dataset(root_dir=img_path, width=IMG_WIDTH, height=IMG_HEIGHT, is_train=True)

    trainloader = DataLoader(
        dataset=dataset,
        batch_size=BATCH_SIZE,
        shuffle=True,
    )

    model = build_model(num_classes=1000)
    model = model.to(device)
    criterion = nn.CrossEntropyLoss()
    optimizer = optim.Adam(model.parameters(), lr=Learning_Rate)
    scheduler = torch.optim.lr_scheduler.ReduceLROnPlateau(
        optimizer, mode='min', factor=0.5, patience=10, min_lr=1e-6
    )

    da_conditions = {
        'Origin': ['Baseline'],
        'Brightness': list(range(0, 210, 10)),
        'ColorTemperature': list(range(0, 210, 10)),
        'Grid_Mask': list(range(0, 110, 10)),
        'Horizontal_Roll': list(range(0, 370, 10))
    }

    da_history = {da: {val: [] for val in vals} for da, vals in da_conditions.items()}

    print("Pre-building all Validation DataLoaders (Level 2)...")
    setup_val_dataloaders(da_conditions)

    start_epoch = 0
    best_loss = float('inf')
    epoch_losses = []

    if os.path.exists(MODEL_PATH):
        try:
            print(f"Loading weights from {MODEL_PATH}")
            checkpoint = torch.load(MODEL_PATH, map_location=device)
            model.load_state_dict(checkpoint['model_state_dict'])
            best_loss = checkpoint['best_loss']
            start_epoch = checkpoint['epoch']
            print("Weight loaded successfully.")
        except Exception as e:
            print(f"Loading failed: {e}, training from scratch.")
    else:
        print("No existing weights found. Training from scratch.")

    print("Start Training...")

    for epoch in range(start_epoch, epochs):
        model.train()
        running_loss = 0.0

        with tqdm(trainloader, desc=f"Epoch {epoch + 1}/{epochs}", ncols=100, leave=True) as loop:
            for img, id_label in loop:
                img, id_label = img.to(device), id_label.to(device)

                optimizer.zero_grad()
                outputs = model(img)
                loss = criterion(outputs, id_label)
                loss.backward()

                torch.nn.utils.clip_grad_norm_(model.parameters(), max_norm=2.0)
                optimizer.step()

                running_loss += loss.item()
                current_lr = optimizer.param_groups[0]['lr']
                loop.set_postfix(loss=f"{loss.item():.4f}")

        avg_loss = running_loss / len(trainloader)
        epoch_losses.append(avg_loss)
        scheduler.step(avg_loss)

        print(f"Epoch {epoch + 1} | Loss: {avg_loss:.4f} | LR: {current_lr:.8f}")

        print(f"Running Inference for 90 DA conditions...")
        for da_type, val_list in da_conditions.items():
            for da_val in val_list:
                val_loader = GLOBAL_VAL_LOADERS[da_type][da_val]
                accuracy = run_inference(model, device, val_loader, da_type, da_val)
                da_history[da_type][da_val].append(accuracy)

        plot_and_save_curves(da_history, epoch)
        save_to_excel(da_history)

        if avg_loss < best_loss:
            best_loss = avg_loss
            checkpoint = {
                'model_state_dict': model.state_dict(),
                'best_loss': best_loss,
                'epoch': epoch + 1
            }
            torch.save(checkpoint, MODEL_PATH)
            print(f"New Best Model Saved to {MODEL_PATH}")

    return epoch_losses


def main():
    Mobilenet_training()


if __name__ == '__main__':
    main()