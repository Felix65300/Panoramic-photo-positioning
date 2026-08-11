import os
import sys
from pathlib import Path

current_dir = Path.cwd()

Project_Root = current_dir.parent

sys.path.append(str(Project_Root))

import pandas as pd
import openpyxl
from copy import copy
from openpyxl.styles import PatternFill, Alignment, Border, Side

FILE_ROOT = Path(Project_Root / 'Figures' / 'EfficientNet_Advanced')

def find_best_epoch(parent_folder):
    file_path = FILE_ROOT / parent_folder / 'DA_Accuracy_Final.xlsx'
    # pd 才能一次取一整欄的資料
    df = pd.read_excel(file_path, header=1, skiprows=[2])
    # wb, sheet 才能取得格式
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    # 取得每個強度的 Best Epoch 並存起來
    best_accuracys = ['Best Epoch']
    for intensity in range(0,201,10):
        best_accuracys.append(df[intensity].idxmax() + 1)
    sheet.append(best_accuracys)

    # 複製格式
    new_row_idx = sheet.max_row
    prev_row_idx = new_row_idx - 1
    # 針對新資料的每個欄位，逐一把上一行的格式複製並套用
    for col in range(1, len(best_accuracys) + 1):
        source_cell = sheet.cell(row=prev_row_idx, column=col) # 舊資料的格子
        target_cell = sheet.cell(row=new_row_idx, column=col)  # 新資料的格子

        # 複製各種格式設定
        target_cell.font = copy(source_cell.font)           # 複製字體
        target_cell.border = copy(source_cell.border)       # 複製 Border 跟 Side (外框線)
        target_cell.fill = copy(source_cell.fill)           # 複製 PatternFill (背景填色)
        target_cell.alignment = copy(source_cell.alignment) # 複製 Alignment (對齊方式)
    wb.save(file_path)

def main():
    categories = ['Baseline', 'DropConnect', 'GeM', 'Layer Freezing']
    for category in categories:
        find_best_epoch(category)
if __name__ == '__main__':
    main()