import os
import sys
from pathlib import Path

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

# data_Step2.py 跨資料夾，所以需要額外動作來輔助 import
# 1. 取得目前檔案的 (Training.py) 所在目錄
current_dir = Path.cwd()

# 2. 取得上一層目錄 (專案的根目錄)
Resnet18 = os.path.dirname(current_dir)
src = os.path.dirname(Resnet18)
Project_Root = os.path.dirname(src)

# 3. 將根目錄加入系統搜尋路徑
sys.path.append(Resnet18)
sys.path.append(src)
sys.path.append(Project_Root)

# 4. 開始 import

import torch
import torch.nn as nn
import torch.optim as optim
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from torch.utils.data import DataLoader
from tqdm import tqdm
from src.data_Step2_5_Test import get_test_dataset
from src.data_Step2_5_Train import get_train_dataset
from src.Resnet18.resnet18_revised_version import get_pano_model
# from src.Custom_model.Convolution_Class import CNN  # ← 這行可刪除

# ---------------------------------
# 1. 設定參數與裝置
# ---------------------------------
BATCH_SIZE = 128 # 根據顯卡記憶體調整 (16 或 32)
Learning_Rate = 1e-4 # Adam 的標準學習率
Num_Epoch = 200
IMG_WIDTH = 224      # ← 改成 224
IMG_HEIGHT = 224     # ← 改成 224
DEVICE = torch.device('cuda')
TRAIN_DIR = Project_Root + '/Datasets/Dataset_Step1'
TEST_ROOT = Project_Root + '/Datasets/Dataset_Step2'
FIG_DIR = Project_Root + '/Figures/Step2.5/Custom_model'
XLSX_DIR = Project_Root + '/Figures/Step2.5/Custom_model'
MODEL_PATH = 'Step2.5_pano_cnn_model.pth'
DA_ACCURACY = {}

TEST_DATALOADER_DICT = {}
TRAIN_DATALOADER = None

def define_data_loaders():
    global TRAIN_DATALOADER, TEST_DATALOADER_DICT, TRAIN_DIR, IMG_WIDTH, IMG_HEIGHT, DEVICE, BATCH_SIZE
    train_dataset = get_train_dataset(TRAIN_DIR, IMG_WIDTH, IMG_HEIGHT, is_train=True)
    TRAIN_DATALOADER = DataLoader(train_dataset
               , batch_size=BATCH_SIZE
               , shuffle=True
               , num_workers=4
               , pin_memory=False)

    da_conditions = {'Brightness': list(range(0, 210, 10)),
                     'Colortemperature': list(range(0, 210, 10)),
                     'Grid_Mask': list(range(0, 110, 10)),
                     'Horizontal_Roll': list(range(0, 370, 10))}
    for category, values in da_conditions.items():
        TEST_DATALOADER_DICT[category] = {}
        DA_ACCURACY[category] = {}
        for val in values:
            DA_ACCURACY[category][val] = list()
            test_dir = TEST_ROOT + f'/{category}/{val}'
            if category == 'Horizontal_Roll':
                test_dir += '°'
            else:
                test_dir += '%'
            test_dataset = get_test_dataset(test_dir)

            TEST_DATALOADER_DICT[category][val] = DataLoader(test_dataset
                                         , batch_size=BATCH_SIZE
                                         , shuffle=False
                                         , num_workers=4
                                         , pin_memory=False)
    TEST_DATALOADER_DICT['Origin'] = {}
    TEST_DATALOADER_DICT['Origin']['Baseline'] = DataLoader(train_dataset
                                                     , batch_size=BATCH_SIZE
                                                     , shuffle=False
                                                     , num_workers=4
                                                     , pin_memory=False)
    DA_ACCURACY['Origin'] = {}
    DA_ACCURACY['Origin']['Baseline'] = list()

def model_test(model):
    global TEST_DATALOADER_DICT, DA_ACCURACY
    model.eval()
    for category, val_dict in TEST_DATALOADER_DICT.items():
        for val, test_loader in val_dict.items():
            correct = 0
            total = 0
            with torch.no_grad():
                for images, labels in tqdm(test_loader, desc="Testing", unit='batch'):
                    images, labels = images.to(DEVICE), labels.to(DEVICE)

                    outputs = model(images)
                    _, predicted = torch.max(outputs, 1)

                    total += labels.size(0)
                    correct += (predicted == labels).sum().item()
                    DA_ACCURACY[category][val].append(100 * correct / total)

def model_training_and_test ():
    global TRAIN_DATALOADER, DEVICE, MODEL_PATH, Num_Epoch
    # ---------------------------------------------------
    # 3. 初始化模型
    # ---------------------------------------------------
    model = get_pano_model(num_classes=len(TRAIN_DATALOADER.dataset.classes)).to(DEVICE)
    # model = CNN(num_classes=len(TRAIN_DATALOADER.dataset.classes)).to(DEVICE)  # ← 這行可刪除
    loss_func = nn.CrossEntropyLoss().to(DEVICE)
    optimizer = optim.Adam(model.parameters(), lr=Learning_Rate)
    scheduler = torch.optim.lr_scheduler.ReduceLROnPlateau(
        optimizer,mode='min', factor=0.5,patience = 10,min_lr = 1e-6
    )

    best_loss = float('inf')

    if os.path.exists(MODEL_PATH):
        checkpoint = torch.load(MODEL_PATH)
        model.load_state_dict(checkpoint['model_state_dict'])
        best_loss = checkpoint['best_loss']

    for epoch in range(Num_Epoch):
        print(f"--> 開始訓練...{epoch+1}/{Num_Epoch}")
        model.train()
        running_loss = 0.0

        with tqdm(TRAIN_DATALOADER, ncols=100) as loop:
            for images, labels in loop:
                images, labels = images.to(DEVICE), labels.to(DEVICE)

                optimizer.zero_grad()
                outputs = model(images)
                loss = loss_func(outputs, labels)
                loss.backward()
                optimizer.step()

                running_loss += loss.item()
                loop.set_postfix(loss=loss.item())
                current_lr = optimizer.param_groups[0]['lr']

        avg_loss = running_loss / len(TRAIN_DATALOADER)
        scheduler.step(avg_loss)

        print(f"Loss: {avg_loss:.4f} | LR: {current_lr:.8f}")

        if avg_loss < best_loss:
            best_loss = avg_loss

            checkpoint = {
                'model_state_dict': model.state_dict(),
                'best_loss': best_loss,
            }

            torch.save(checkpoint, MODEL_PATH)
        model_test(model)

def generate_figure():
    global DA_ACCURACY,FIG_DIR
    epochs = np.arange(1, 201)
    plt.rcParams.update({
        'font.family': 'serif',
        'font.serif': ['Times New Roman'],
        'font.size': 10,
        'axes.labelsize': 10,
        'axes.titlesize': 10,
        'xtick.labelsize': 8,
        'ytick.labelsize': 8,
        'legend.fontsize': 8,
        'legend.frameon': False,
        'lines.linewidth': 1.2,
        'axes.linewidth': 0.8,
        'grid.linestyle': '--',
        'grid.alpha': 0.5,
        'figure.dpi': 300
    })
    color = '#d62728'
    for category, values in DA_ACCURACY.items():
        for val,accuracy in values.items():
            fig, ax = plt.subplots(figsize=(3.5, 2.5))
            ax.plot(epochs, accuracy, label="Model", color=color, linestyle='-')
            ax.set_xlabel("Epochs")
            ax.set_ylabel("Accuracy (%)")
            ax.legend(loc='lower right')
            ax.set_ylim(0, 100)
            plt.tight_layout()
            fig_name = f"{category}"
            if category == 'Horizontal_Roll':
                fig_name += f"_{val}°"
            elif category != 'Origin':
                fig_name += f"_{val}%"
            fig_name += '_Accuracy.png'
            save_path = Path(FIG_DIR) / f"{category}" / f"{fig_name}"
            save_path.parent.mkdir(parents=True, exist_ok=True)
            plt.savefig(save_path, bbox_inches='tight')
            plt.close()

def gernerate_xlsx():
    global DA_ACCURACY
    formatted_data = {}
    for category, values in DA_ACCURACY.items():
        for val,accuracy in values.items():
            formatted_data[(category, val)] = accuracy
    df = pd.DataFrame(formatted_data)
    df.index = range(1, len(df) + 1)
    df.index.name = 'Epoch'
    df.columns.names = ['DA Topic', 'Intensity']
    file_path = Path(XLSX_DIR) / "DA_Accuracy_Final.xlsx"
    df.to_excel(file_path, engine='openpyxl')

    font_header = Font(name="Times New Roman", size=11, bold=True, color="FFFFFF")
    font_index = Font(name="Times New Roman", size=11, bold=True)
    font_regular = Font(name="Times New Roman", size=11)
    fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F6F9", end_color="F2F6F9", fill_type="solid")
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    thin_side = Side(border_style='thin', color='D3D3D3')
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    ws.title = "Training Logs"
    max_row = ws.max_row
    max_col = ws.max_column

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border_all
            if r <= 3:
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
            else:
                cell.font = font_regular
                if c == 1:
                    cell.font = font_index
                    cell.alignment = align_center
                else:
                    cell.alignment = align_right
                    cell.number_format = "0.00"
                if r % 2 == 0:
                    cell.fill = fill_zebra
    ws.freeze_panes = "B4"
    ws.column_dimensions['A'].width = 10
    for col_idx in range(2, max_col + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 14
    wb.save(file_path)

def main():
    define_data_loaders()
    model_training_and_test()
    generate_figure()
    gernerate_xlsx()

if __name__ == '__main__':
    main()